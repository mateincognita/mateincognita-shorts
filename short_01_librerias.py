
# Short #01 - mateIncognita
# Las 4 librerias esenciales para analisis de datos
# github.com/mateincognita/mateincognita-shorts


# 1. PANDAS 
# La libreria mas usada para manipular datos
# Lee Excel, CSV, JSON y mas

import pandas as pd

df = pd.read_excel("data.xlsx")   # leer archivo
df.head()                          # ver primeras 5 filas
df.describe()                      # estadisticas basicas
df.groupby("columna").sum()        # agrupar y sumar


# 2. PLOTLY 
# Graficas interactivas en una sola linea
# Hover, zoom, descarga incluidos

import plotly.express as px

fig = px.bar(df, x="categoria", y="valor",
             title="Mi grafica",
             color="categoria")
fig.show()                         # abre en el navegador
fig.write_image("grafica.png")     # guarda como imagen


# 3. OPENPYXL 
# Lee y escribe archivos Excel directamente
# Sin necesidad de tener Excel instalado
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

wb = load_workbook("data.xlsx")    # abrir archivo
ws = wb.active                     # hoja activa

ws["A1"] = "Actualizado"           # editar celda
wb.save("data_editado.xlsx")       # guardar cambios


# 4. REPORTLAB 
# Genera reportes PDF automaticamente
# Ideal para reportes periodicos sin intervension manual
from reportlab.platypus import SimpleDocTemplate, Table
from reportlab.lib.pagesizes import letter

doc = SimpleDocTemplate("reporte.pdf", pagesize=letter)
elementos = []
doc.build(elementos)               # genera el PDF


# INSTALACION 
# pip install pandas openpyxl plotly kaleido reportlab



